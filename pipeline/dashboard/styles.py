"""Shared Excel styling constants and helper functions."""
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd

# ---------------------------------------------------------------------------
# Color palette
# ---------------------------------------------------------------------------
NAVY = '1F4E79'
WHITE = 'FFFFFF'
LIGHT_GRAY = 'F5F5F5'
MEDIUM_GRAY = 'E0E0E0'

# Tier colors (text)
TIER_FONT = {
    'LOW':      Font(color='059669', bold=True),
    'MODERATE': Font(color='D97706', bold=True),
    'HIGH':     Font(color='DC2626', bold=True),
    'CRITICAL': Font(color='991B1B', bold=True),
}

# Tier colors (cell fill — soft background)
TIER_FILL = {
    'LOW':      PatternFill('solid', fgColor='ECFDF5'),
    'MODERATE': PatternFill('solid', fgColor='FFFBEB'),
    'HIGH':     PatternFill('solid', fgColor='FEF2F2'),
    'CRITICAL': PatternFill('solid', fgColor='FEE2E2'),
}

# Reusable styles
HEADER_FONT = Font(bold=True, color=WHITE, size=10)
HEADER_FILL = PatternFill('solid', fgColor=NAVY)
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)

BODY_ALIGN = Alignment(vertical='center')
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
RIGHT_ALIGN = Alignment(horizontal='right', vertical='center')

TITLE_FONT = Font(bold=True, size=14, color=NAVY)
SUBTITLE_FONT = Font(bold=True, size=11, color=NAVY)
SECTION_FONT = Font(bold=True, size=10, color='333333')

THIN_BORDER = Border(
    bottom=Side(style='thin', color=MEDIUM_GRAY),
)

ALT_ROW_FILL = PatternFill('solid', fgColor=LIGHT_GRAY)


# ---------------------------------------------------------------------------
# Helper functions
# ---------------------------------------------------------------------------

def write_table(ws, start_row, start_col, df, number_formats=None):
    """Write a DataFrame as a styled table.

    Args:
        ws: openpyxl worksheet
        start_row: 1-based row for the header
        start_col: 1-based column offset
        df: pandas DataFrame
        number_formats: optional dict {col_name: excel_format_string}
    """
    if number_formats is None:
        number_formats = {}

    # Header
    for j, col in enumerate(df.columns):
        cell = ws.cell(row=start_row, column=start_col + j, value=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN

    # Body
    for i in range(len(df)):
        for j, col in enumerate(df.columns):
            val = df.iloc[i, j]
            # Convert numpy types to native Python
            if hasattr(val, 'item'):
                val = val.item()
            if pd.isna(val):
                val = None
            cell = ws.cell(row=start_row + 1 + i, column=start_col + j, value=val)
            cell.alignment = BODY_ALIGN
            cell.border = THIN_BORDER
            if col in number_formats:
                cell.number_format = number_formats[col]
            # Alternating row shading
            if i % 2 == 1:
                cell.fill = ALT_ROW_FILL


def auto_width(ws, min_width=10, max_width=55):
    """Auto-fit column widths, clamped between min and max."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(min_width, max_len + 3), max_width)


def write_kpi(ws, row, col, label, value, value_font=None):
    """Write a KPI card: label above, bold value below."""
    label_cell = ws.cell(row=row, column=col, value=label)
    label_cell.font = Font(size=8, color='888888', bold=True)
    label_cell.alignment = CENTER_ALIGN

    val_cell = ws.cell(row=row + 1, column=col, value=value)
    val_cell.font = value_font or Font(size=16, bold=True, color=NAVY)
    val_cell.alignment = CENTER_ALIGN


def write_section_header(ws, row, col, text, span=1):
    """Write a section header spanning multiple columns."""
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = SUBTITLE_FONT
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + span - 1)


def apply_tier_formatting(ws, tier_col_idx, start_row, end_row):
    """Color the tier column cells based on tier value."""
    for row in range(start_row, end_row + 1):
        cell = ws.cell(row=row, column=tier_col_idx)
        tier = str(cell.value or '').upper()
        if tier in TIER_FONT:
            cell.font = TIER_FONT[tier]
        if tier in TIER_FILL:
            cell.fill = TIER_FILL[tier]
        cell.alignment = CENTER_ALIGN


def apply_row_tier_shading(ws, tier_col_idx, start_row, end_row, num_cols, col_offset=1):
    """Apply soft tier background to entire rows."""
    for row in range(start_row, end_row + 1):
        tier = str(ws.cell(row=row, column=tier_col_idx).value or '').upper()
        if tier in TIER_FILL:
            for c in range(col_offset, col_offset + num_cols):
                ws.cell(row=row, column=c).fill = TIER_FILL[tier]
