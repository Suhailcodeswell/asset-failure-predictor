"""Build the OpsInsight Risk Dashboard Excel workbook."""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .data_loader import DashboardData
from .styles import (
    NAVY, WHITE, HEADER_FONT, HEADER_FILL, HEADER_ALIGN,
    TITLE_FONT, SUBTITLE_FONT, SECTION_FONT,
    TIER_FONT, TIER_FILL, CENTER_ALIGN, RIGHT_ALIGN,
    write_table, auto_width, write_kpi, write_section_header,
    apply_tier_formatting, apply_row_tier_shading,
)

TIER_ORDER = ['LOW', 'MODERATE', 'HIGH', 'CRITICAL']


class ExcelDashboardBuilder:
    """Generates a multi-sheet Excel dashboard from scored fleet data."""

    def __init__(self, data: DashboardData, scored_fleet: list[dict],
                 fleet_summary: dict, pm_buckets: list[dict],
                 model_analysis: dict, component_comparison: list[dict]):
        self.data = data
        self.scored = scored_fleet
        self.summary = fleet_summary
        self.pm_buckets = pm_buckets
        self.model_analysis = model_analysis
        self.comp_comparison = component_comparison

    def build(self, output_path: str) -> str:
        """Build and save the workbook. Returns the output path."""
        wb = Workbook()

        # Sheet 1: Fleet Summary (rename default sheet)
        ws1 = wb.active
        ws1.title = 'Fleet Summary'
        self._build_fleet_summary(ws1)

        # Sheet 2: Fleet Risk Ranking
        ws2 = wb.create_sheet('Fleet Risk Ranking')
        self._build_fleet_ranking(ws2)

        # Sheet 3: Unit Lookup
        ws3 = wb.create_sheet('Unit Lookup')
        self._build_unit_lookup(ws3)

        # Sheet 4: Unit Features
        ws4 = wb.create_sheet('Unit Features')
        self._build_unit_features(ws4)

        # Sheet 5: Model Comparison
        ws5 = wb.create_sheet('Model Comparison')
        self._build_model_comparison(ws5)

        # Sheet 6: Component Analysis
        ws6 = wb.create_sheet('Component Analysis')
        self._build_component_analysis(ws6)

        # Sheet 7: _Data (hidden)
        ws7 = wb.create_sheet('_Data')
        self._build_data_sheet(ws7)
        ws7.sheet_state = 'hidden'

        wb.save(output_path)
        print(f'  Excel: {output_path}')
        return output_path

    # ==================================================================
    # Sheet 1: Fleet Summary
    # ==================================================================
    def _build_fleet_summary(self, ws):
        s = self.summary

        # Title
        ws.cell(row=1, column=1, value='OpsInsight Risk Dashboard — Fleet Summary').font = TITLE_FONT
        ws.cell(row=2, column=1, value='Reefer Fleet Intelligence').font = Font(color='888888', size=10)

        # KPI row
        kpis = [
            ('TOTAL UNITS', s['total_units'], None),
            ('ACTIVE', s['active_units'], None),
            ('CRITICAL', s['tier_counts'].get('CRITICAL', 0), Font(size=16, bold=True, color='991B1B')),
            ('HIGH', s['tier_counts'].get('HIGH', 0), Font(size=16, bold=True, color='DC2626')),
            ('SHUTDOWNS', s['total_shutdowns'], None),
            ('FLEET SD RATE', f"{s['fleet_sd_rate']}%", None),
        ]
        for j, (label, value, vfont) in enumerate(kpis):
            write_kpi(ws, 4, j + 1, label, value, vfont)

        # Tier distribution table
        write_section_header(ws, 7, 1, 'Risk Tier Distribution', 5)
        tier_df = pd.DataFrame([
            {'Tier': t, 'Count': s['tier_counts'].get(t, 0),
             '% of Fleet': f"{s['tier_counts'].get(t, 0) / max(s['total_units'], 1) * 100:.1f}%"}
            for t in TIER_ORDER
        ])
        write_table(ws, 8, 1, tier_df)
        apply_tier_formatting(ws, 1, 9, 9 + len(TIER_ORDER) - 1)

        # Make Comparison (SLXi vs Standard)
        row_start = 14
        write_section_header(ws, row_start, 1, 'SLXi vs Standard Comparison', 6)

        slxi_units = [u for u in self.scored if u['is_slxi']]
        std_units = [u for u in self.scored if not u['is_slxi']]

        def _stats(units):
            n = len(units)
            if n == 0:
                return {'Count': 0, 'Avg Risk %': '-', 'SD Rate %': '-', 'PM Overdue': 0, 'Avg Repair $': '-'}
            sd = sum(u['shutdown_count'] for u in units)
            trips = sum(u['total_rail_legs'] for u in units)
            return {
                'Count': n,
                'Avg Risk %': round(sum(u['risk_score'] for u in units) / n * 100, 2),
                'SD Rate %': round(sd / max(trips, 1) * 100, 2),
                'PM Overdue': sum(1 for u in units if u['is_pm_overdue']),
                'Avg Repair $': round(sum(u['nonpm_cum_cost'] for u in units) / n),
            }

        make_rows = [{'Type': 'SLXi', **_stats(slxi_units)},
                     {'Type': 'Standard', **_stats(std_units)}]
        make_df = pd.DataFrame(make_rows)
        write_table(ws, row_start + 1, 1, make_df)

        # Model Breakdown
        row_start = row_start + len(make_rows) + 3
        write_section_header(ws, row_start, 1, 'Model Breakdown', 7)

        by_model = {}
        for u in self.scored:
            m = u['reefer_model'] or 'Unknown'
            if m == 'nan':
                m = 'Unknown'
            by_model.setdefault(m, []).append(u)

        model_rows = []
        for model, units in sorted(by_model.items(), key=lambda x: -len(x[1])):
            n = len(units)
            if n < 2:
                continue
            sd = sum(u['shutdown_count'] for u in units)
            trips = sum(u['total_rail_legs'] for u in units)
            elevated = sum(1 for u in units if u['tier'] in ('HIGH', 'CRITICAL'))
            model_rows.append({
                'Model': model,
                'Count': n,
                'Avg Risk %': round(sum(u['risk_score'] for u in units) / n * 100, 2),
                'SD Rate %': round(sd / max(trips, 1) * 100, 2),
                'HIGH+': elevated,
                'PM Overdue': sum(1 for u in units if u['is_pm_overdue']),
                'Avg Repair $': round(sum(u['nonpm_cum_cost'] for u in units) / n),
            })
        model_df = pd.DataFrame(model_rows)
        write_table(ws, row_start + 1, 1, model_df)

        # PM Delay analysis
        row_start = row_start + len(model_rows) + 3
        write_section_header(ws, row_start, 1, 'PM Delay vs Risk & Shutdown Rate', 4)
        pm_df = pd.DataFrame(self.pm_buckets)
        pm_df.columns = ['PM Status', 'Units', 'Avg Risk %', 'Shutdown Rate %']
        write_table(ws, row_start + 1, 1, pm_df)

        auto_width(ws)

    # ==================================================================
    # Sheet 2: Fleet Risk Ranking
    # ==================================================================
    def _build_fleet_ranking(self, ws):
        ws.cell(row=1, column=1, value='Fleet Risk Ranking').font = TITLE_FONT
        ws.cell(row=2, column=1, value='All units ranked by shutdown risk (latest trip features)').font = Font(color='888888', size=9)

        cols = ['Rank', 'Unit', 'Trailer', 'Model', 'Year', 'Type', 'Risk %',
                'Tier', 'Shutdowns', 'Trips', 'PM Status', 'Repair $', 'Top Risk Factor']
        fmts = {'Risk %': '0.00%', 'Repair $': '$#,##0'}

        rows = []
        for s in self.scored:
            top_factor = s['risk_factors'][0]['label'] if s['risk_factors'] else ''
            pm_status = f"{s['days_pm_overdue']}d overdue" if s['is_pm_overdue'] else 'OK'
            rows.append({
                'Rank': s['rank'],
                'Unit': s['unit'],
                'Trailer': s['trailer'],
                'Model': s['reefer_model'],
                'Year': s['reefer_year'],
                'Type': 'SLXi' if s['is_slxi'] else 'Standard',
                'Risk %': s['risk_score'],
                'Tier': s['tier'],
                'Shutdowns': s['shutdown_count'],
                'Trips': s['total_rail_legs'],
                'PM Status': pm_status,
                'Repair $': s['nonpm_cum_cost'],
                'Top Risk Factor': top_factor,
            })

        df = pd.DataFrame(rows)
        write_table(ws, 4, 1, df, number_formats=fmts)

        # Tier formatting on column 8 (Tier), row shading
        tier_col = 8
        data_start = 5
        data_end = 4 + len(rows)
        apply_tier_formatting(ws, tier_col, data_start, data_end)
        apply_row_tier_shading(ws, tier_col, data_start, data_end, len(cols))

        # Risk % column formatting
        risk_col = 7
        for r in range(data_start, data_end + 1):
            ws.cell(row=r, column=risk_col).number_format = '0.00%'

        auto_width(ws)
        ws.freeze_panes = 'A5'

    # ==================================================================
    # Sheet 3: Unit Lookup
    # ==================================================================
    def _build_unit_lookup(self, ws):
        n = len(self.scored)
        ws.cell(row=1, column=1, value='Unit Lookup').font = TITLE_FONT
        ws.cell(row=2, column=1, value='Select a unit to view its risk profile and SHAP-based risk factors').font = Font(color='888888', size=9)

        # Unit selector
        ws.cell(row=4, column=1, value='SELECT UNIT:').font = SECTION_FONT
        ws.cell(row=4, column=2, value=self.scored[0]['unit'])  # default
        dv = DataValidation(type='list', formula1=f"=_Data!$A$2:$A${n + 1}", allow_blank=False)
        ws.add_data_validation(dv)
        dv.add(ws['B4'])

        # INDEX/MATCH helper — column letter in _Data by offset
        def match_formula(data_col_letter):
            return f'=IFERROR(INDEX(_Data!${data_col_letter}:${data_col_letter},MATCH($B$4,_Data!$A:$A,0)),"")'

        # Risk Profile section
        write_section_header(ws, 6, 1, 'Risk Profile', 3)
        labels_left = [
            ('Risk Score', 'B'),   # col B = risk_score
            ('Risk %', 'C'),       # col C = risk_percent
            ('Tier', 'D'),         # col D = tier
        ]
        labels_right = [
            ('Model', 'E'),
            ('Year', 'F'),
            ('Make', 'G'),
        ]
        for i, (label, col_letter) in enumerate(labels_left):
            ws.cell(row=7 + i, column=1, value=label).font = Font(bold=True, color='555555')
            ws.cell(row=7 + i, column=2, value=match_formula(col_letter))

        for i, (label, col_letter) in enumerate(labels_right):
            ws.cell(row=7 + i, column=3, value=label).font = Font(bold=True, color='555555')
            ws.cell(row=7 + i, column=4, value=match_formula(col_letter))

        # Operational stats
        write_section_header(ws, 11, 1, 'Operational History', 4)
        op_labels = [
            ('Shutdowns', 'H'),
            ('Total Trips', 'I'),
            ('Days Since PM', 'J'),
            ('PM Overdue (days)', 'K'),
            ('Repair Cost (cumul.)', 'L'),
            ('Repairs in 90d', 'M'),
            ('Unique Comp Codes', 'N'),
        ]
        for i, (label, col_letter) in enumerate(op_labels):
            ws.cell(row=12 + i, column=1, value=label).font = Font(bold=True, color='555555')
            ws.cell(row=12 + i, column=2, value=match_formula(col_letter))

        # SHAP Risk Factors section
        write_section_header(ws, 20, 1, 'Top Risk Factors (SHAP)', 4)
        ws.cell(row=21, column=1, value='Factor').font = HEADER_FONT
        ws.cell(row=21, column=1).fill = HEADER_FILL
        ws.cell(row=21, column=2, value='Value').font = HEADER_FONT
        ws.cell(row=21, column=2).fill = HEADER_FILL
        ws.cell(row=21, column=3, value='Impact').font = HEADER_FONT
        ws.cell(row=21, column=3).fill = HEADER_FILL
        ws.cell(row=21, column=4, value='Direction').font = HEADER_FONT
        ws.cell(row=21, column=4).fill = HEADER_FILL

        # Factors stored in _Data cols O onward (5 factors x 4 fields each = 20 cols)
        # O=label, P=value, Q=shap_value, R=direction, S=label2, ...
        factor_start_col = ord('O') - ord('A')  # 0-based: 14
        for fi in range(5):
            base = factor_start_col + fi * 4
            label_col = get_column_letter(base + 1)
            value_col = get_column_letter(base + 2)
            shap_col = get_column_letter(base + 3)
            dir_col = get_column_letter(base + 4)
            r = 22 + fi
            ws.cell(row=r, column=1, value=match_formula(label_col))
            ws.cell(row=r, column=2, value=match_formula(value_col))
            ws.cell(row=r, column=3, value=match_formula(shap_col))
            ws.cell(row=r, column=4, value=match_formula(dir_col))

        # Recommendations section
        write_section_header(ws, 28, 1, 'Recommendations', 4)
        # Recs start at column 35 in _Data (14 base + 20 SHAP + 1)
        for ri in range(3):
            r = 29 + ri
            action_col = get_column_letter(35 + ri * 2)
            urgency_col = get_column_letter(36 + ri * 2)
            ws.cell(row=r, column=1, value=match_formula(urgency_col)).font = Font(bold=True)
            ws.cell(row=r, column=2, value=match_formula(action_col))

        auto_width(ws)

    # ==================================================================
    # Sheet 4: Unit Features (all model features for selected unit)
    # ==================================================================
    def _build_unit_features(self, ws):
        n = len(self.scored)
        ws.cell(row=1, column=1, value='Unit Feature Reference').font = TITLE_FONT
        ws.cell(row=2, column=1,
                value='All model input features for the selected unit vs fleet median. '
                      'For live prediction, use the webapp.').font = Font(color='888888', size=9)

        # Unit selector (same dropdown)
        ws.cell(row=4, column=1, value='SELECT UNIT:').font = SECTION_FONT
        ws.cell(row=4, column=2, value=self.scored[0]['unit'])
        dv = DataValidation(type='list', formula1=f"=_Data!$A$2:$A${n + 1}", allow_blank=False)
        ws.add_data_validation(dv)
        dv.add(ws['B4'])

        # Feature table header
        headers = ['#', 'Category', 'Feature', 'Label', "This Unit's Value", 'Fleet Median']
        header_row = 6
        for j, h in enumerate(headers):
            cell = ws.cell(row=header_row, column=j + 1, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER_ALIGN

        # Build feature rows with INDEX/MATCH to _Data
        # Features start at column 41 (14 base + 20 SHAP + 6 recs + 1)
        feat_data_start_col = 41
        data = self.data

        for i, feat in enumerate(data.feature_cols):
            r = header_row + 1 + i
            label = data.feature_labels.get(feat, feat.replace('_', ' ').title())
            category = data.feature_categories.get(feat, 'Other')
            median = data.feature_defaults.get(feat, 0)

            ws.cell(row=r, column=1, value=i + 1).alignment = CENTER_ALIGN
            ws.cell(row=r, column=2, value=category)
            ws.cell(row=r, column=3, value=feat).font = Font(size=9, color='666666')
            ws.cell(row=r, column=4, value=label)

            # Dynamic value via INDEX/MATCH
            feat_col_letter = get_column_letter(feat_data_start_col + i)
            formula = f'=IFERROR(INDEX(_Data!${feat_col_letter}:${feat_col_letter},MATCH($B$4,_Data!$A:$A,0)),"")'
            ws.cell(row=r, column=5, value=formula)

            # Fleet median (static)
            if isinstance(median, float):
                ws.cell(row=r, column=6, value=round(median, 4))
            else:
                ws.cell(row=r, column=6, value=median)

        auto_width(ws)
        ws.freeze_panes = 'A7'

    # ==================================================================
    # Sheet 5: Model Comparison
    # ==================================================================
    def _build_model_comparison(self, ws):
        ws.cell(row=1, column=1, value='Model Comparison').font = TITLE_FONT
        ws.cell(row=2, column=1, value='Reefer model performance comparison and cohort analysis').font = Font(color='888888', size=9)

        # Model summary table
        write_section_header(ws, 4, 1, 'Model Performance Summary', 9)
        model_rows = []
        for m in self.model_analysis['models']:
            model_rows.append({
                'Model': m['model'],
                'Units': m['count'],
                'SD Rate %': m['sd_rate'],
                'Avg Risk %': m['avg_risk'],
                'Avg Repair $': m['avg_repair'],
                'Cost/Trip': m['cost_per_trip'],
                'PM Overdue %': m['pm_overdue_pct'],
                'Elevated %': m['elevated_pct'],
                'Shutdowns': m['shutdowns'],
            })
        if model_rows:
            model_df = pd.DataFrame(model_rows)
            write_table(ws, 5, 1, model_df, number_formats={'Cost/Trip': '$#,##0.00', 'Avg Repair $': '$#,##0'})

        # Worst cohorts table
        cohort_row = 5 + len(model_rows) + 3
        write_section_header(ws, cohort_row, 1, 'Worst Performing Cohorts', 8)
        cohort_rows = []
        for c in self.model_analysis['worst_cohorts']:
            cohort_rows.append({
                'Model': c['model'],
                'Year': c['year'],
                'Age': c['age'],
                'Units': c['count'],
                'SD Rate %': c['sd_rate'],
                'Avg Risk %': c['avg_risk'],
                'Avg Repair $': c['avg_repair'],
                'Elevated': c['elevated'],
            })
        if cohort_rows:
            cohort_df = pd.DataFrame(cohort_rows)
            write_table(ws, cohort_row + 1, 1, cohort_df, number_formats={'Avg Repair $': '$#,##0'})

        # Component vulnerability heatmap
        heat_row = cohort_row + len(cohort_rows) + 4
        write_section_header(ws, heat_row, 1, 'Component Vulnerability by Model', 9)

        comp_labels = ['Compressor', 'Evaporator', 'Diagnosis', 'Electrical',
                       'Fuel System', 'Starter/Cranking', 'Condenser', 'Engine']
        heat_headers = ['Model'] + comp_labels
        for j, h in enumerate(heat_headers):
            cell = ws.cell(row=heat_row + 1, column=j + 1, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGN

        for i, m in enumerate(self.model_analysis['models']):
            r = heat_row + 2 + i
            ws.cell(row=r, column=1, value=m['model']).font = Font(bold=True)
            for j, comp in enumerate(comp_labels):
                val = m['components'].get(comp, 0)
                cell = ws.cell(row=r, column=j + 2, value=val)
                cell.alignment = CENTER_ALIGN
                # Heat color: white -> yellow -> red
                if val >= 2:
                    cell.fill = PatternFill('solid', fgColor='FEE2E2')
                    cell.font = Font(bold=True, color='991B1B')
                elif val >= 1:
                    cell.fill = PatternFill('solid', fgColor='FEF3C7')
                    cell.font = Font(color='92400E')
                elif val >= 0.5:
                    cell.fill = PatternFill('solid', fgColor='FFF7ED')

        auto_width(ws)

    # ==================================================================
    # Sheet 6: Component Analysis
    # ==================================================================
    def _build_component_analysis(self, ws):
        ws.cell(row=1, column=1, value='Component Analysis — SLXi vs Standard').font = TITLE_FONT
        ws.cell(row=2, column=1, value='Average cumulative repairs per unit by component type').font = Font(color='888888', size=9)

        slxi = [u for u in self.scored if u['is_slxi']]
        std = [u for u in self.scored if not u['is_slxi']]
        ws.cell(row=3, column=1,
                value=f'SLXi: {len(slxi)} units  |  Standard: {len(std)} units').font = Font(color='555555', size=9)

        comp_df = pd.DataFrame(self.comp_comparison)
        comp_df.columns = ['Component', 'SLXi Avg', 'Standard Avg', 'Delta']
        write_table(ws, 5, 1, comp_df, number_formats={'SLXi Avg': '0.000', 'Standard Avg': '0.000', 'Delta': '+0.000;-0.000;0.000'})

        # Color delta column
        for i in range(len(self.comp_comparison)):
            cell = ws.cell(row=6 + i, column=4)
            delta = self.comp_comparison[i]['delta']
            if delta > 0:
                cell.font = Font(bold=True, color='DC2626')
            elif delta < 0:
                cell.font = Font(bold=True, color='059669')
            else:
                cell.font = Font(color='999999')

        auto_width(ws)

    # ==================================================================
    # Sheet 7: _Data (hidden reference sheet for formulas)
    # ==================================================================
    def _build_data_sheet(self, ws):
        data = self.data
        n = len(self.scored)

        # Column layout:
        # A: unit, B: risk_score, C: risk_percent, D: tier, E: model, F: year, G: type
        # H: shutdowns, I: trips, J: days_since_pm, K: days_pm_overdue, L: repair_cum_cost
        # M: repair_count_90d, N: unique_compcodes_cum
        # O-AH: 5 SHAP factors (label, value, shap_value, direction) = 20 cols
        # AI-AN: 3 recommendations (action, urgency) = 6 cols
        # AO onward: feature values (83 as of 2026-04-09)

        # Header
        base_headers = [
            'unit', 'risk_score', 'risk_percent', 'tier', 'model', 'year', 'make',
            'shutdowns', 'trips', 'days_since_pm', 'days_pm_overdue',
            'repair_cum_cost', 'repair_count_90d', 'unique_compcodes_cum',
        ]
        # SHAP factor headers (5 factors x 4 fields)
        for fi in range(5):
            base_headers += [f'factor{fi+1}_label', f'factor{fi+1}_value',
                             f'factor{fi+1}_shap', f'factor{fi+1}_direction']
        # Recommendation headers (3 recs x 2 fields)
        for ri in range(3):
            base_headers += [f'rec{ri+1}_action', f'rec{ri+1}_urgency']
        # Feature headers
        base_headers += data.feature_cols

        for j, h in enumerate(base_headers):
            cell = ws.cell(row=1, column=j + 1, value=h)
            cell.font = Font(bold=True, size=8, color='555555')

        # Data rows
        for i, s in enumerate(self.scored):
            r = i + 2

            # Base fields
            ws.cell(row=r, column=1, value=s['unit']).number_format = '@'
            ws.cell(row=r, column=2, value=s['risk_score'])
            ws.cell(row=r, column=3, value=s['risk_percent'])
            ws.cell(row=r, column=4, value=s['tier'])
            ws.cell(row=r, column=5, value=s['reefer_model'])
            ws.cell(row=r, column=6, value=s['reefer_year'])
            ws.cell(row=r, column=7, value='SLXi' if s['is_slxi'] else 'Standard')
            ws.cell(row=r, column=8, value=s['shutdown_count'])
            ws.cell(row=r, column=9, value=s['total_rail_legs'])
            ws.cell(row=r, column=10, value=s['days_since_last_pm'])
            ws.cell(row=r, column=11, value=s['days_pm_overdue'])
            ws.cell(row=r, column=12, value=s['repair_cum_cost'])
            ws.cell(row=r, column=13, value=s['repair_count_90d'])
            ws.cell(row=r, column=14, value=s['unique_compcodes_cum'])

            # SHAP factors (5 x 4: label, value, shap_value, direction)
            col_offset = 15
            for fi in range(5):
                if fi < len(s['risk_factors']):
                    f = s['risk_factors'][fi]
                    ws.cell(row=r, column=col_offset + fi * 4, value=f['label'])
                    val = f['value']
                    if isinstance(val, float):
                        val = round(val, 2)
                    ws.cell(row=r, column=col_offset + fi * 4 + 1, value=str(val) if val is not None else '')
                    ws.cell(row=r, column=col_offset + fi * 4 + 2, value=round(f['shap_value'], 4))
                    ws.cell(row=r, column=col_offset + fi * 4 + 3, value=f['direction'])

            # Recommendations (3 x 2)
            rec_offset = col_offset + 20  # 35
            for ri in range(3):
                if ri < len(s['recommendations']):
                    rec = s['recommendations'][ri]
                    ws.cell(row=r, column=rec_offset + ri * 2, value=rec['action'])
                    ws.cell(row=r, column=rec_offset + ri * 2 + 1, value=rec['urgency'])

            # Feature values (n_features cols)
            feat_offset = rec_offset + 6  # 41
            for j, feat in enumerate(data.feature_cols):
                val = s['feature_values'].get(feat, 0)
                if isinstance(val, float):
                    val = round(val, 6)
                ws.cell(row=r, column=feat_offset + j, value=val)
